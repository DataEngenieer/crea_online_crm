from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import pandas as pd
import logging

from .models import Cliente, Gestion, AcuerdoPago, CuotaAcuerdo

# Diccionario con los campos disponibles para cada tipo de reporte
from django.contrib.auth.models import User, Group

CAMPOS_DISPONIBLES = {
    'clientes': [
        ('documento', 'Documento'),
        ('nombre_completo', 'Nombre Completo'),
        ('deuda_total', 'Deuda Total'),
        ('total_dias_mora', 'Días de Mora'),
        ('telefono_celular', 'Teléfono'),
        ('email', 'Email'),
        ('estado', 'Estado'),
        ('fecha_registro', 'Fecha de Registro'),
        ('ultima_gestion', 'Última Gestión Realizada'),
    ],
    'gestiones': [
        ('fecha_hora_gestion', 'Fecha y Hora'),
        ('canal_contacto', 'Canal'),
        ('cliente__documento', 'Documento Cliente'),
        ('cliente__nombre_completo', 'Nombre Cliente'),
        ('cliente__referencia', 'Referencia Cliente'),
        ('cliente__telefono_celular', 'Teléfono Cliente'),
        ('cliente__email', 'Email Cliente'),
        ('cliente__deuda_total', 'Deuda Total Cliente'),
        ('cliente__total_dias_mora', 'Días Mora Cliente'),
        ('cliente__ciudad', 'Ciudad Cliente'),
        ('cliente__direccion_1', 'Dirección Cliente'),
        ('cliente__observaciones_adicionales', 'Observaciones Cliente'),
        ('referencia_producto', 'Referencia Producto'),
        ('comprobante_pago', 'Comprobante de Pago'),
        ('tipo_gestion_n1', 'Tipo Gestión N1'),
        ('estado_contacto', 'Estado Contacto'),
        ('usuario_gestion__username', 'Usuario'),
        ('comentarios', 'Comentarios'),
        ('acuerdos__monto_total', 'Monto Acuerdo'),
        ('acuerdos__numero_cuotas', 'Número Cuotas'),
        ('acuerdos__fecha_creacion', 'Fecha Acuerdo'),
        ('acuerdos__estado', 'Estado Acuerdo'),
    ],
    'usuarios': [
        ('username', 'Usuario'),
        ('first_name', 'Nombre'),
        ('last_name', 'Apellido'),
        ('email', 'Email'),
        ('is_active', 'Activo'),
        ('date_joined', 'Fecha Registro'),
        ('last_login', 'Último Acceso'),
        ('groups', 'Grupos'),
    ],
    'acuerdos': [
        ('id', 'ID Acuerdo'),
        ('cliente__documento', 'Documento Cliente'),
        ('cliente__nombre_completo', 'Nombre Cliente'),
        ('cliente__ciudad', 'Ciudad Cliente'),
        ('cliente__email', 'Email Cliente'),
        ('cliente__telefono_celular', 'Teléfono Cliente'),
        ('cliente__referencia', 'Referencia Cliente'),
        ('cliente__direccion_1', 'Dirección Cliente'),
        ('cliente__observaciones_adicionales', 'Observaciones Cliente'),
        ('monto_total', 'Monto Total'),
        ('numero_cuotas', 'Número de Cuotas'),
        ('cuotas_pagadas', 'Cuotas Pagadas'),
        ('monto_pagado', 'Monto Total Pagado'),
        ('fecha_ultimo_pago', 'Fecha Último Pago'),
        ('estado_cuotas', 'Estado Global Cuotas'),
        ('comprobantes_pago', 'Comprobantes de Pago'),
        ('fecha_ultima_gestion', 'Fecha Última Gestión'),
        ('usuario_ultima_gestion', 'Usuario Última Gestión'),
        ('observaciones_ultima_gestion', 'Observaciones Última Gestión'),
        ('estado', 'Estado'),
        ('fecha_creacion', 'Fecha de Creación'),
        ('fecha_vencimiento', 'Fecha de Vencimiento'),
    ],
    'pagos': [
        ('acuerdo__cliente__documento', 'Documento Cliente'),
        ('acuerdo__cliente__nombre_completo', 'Nombre Cliente'),
        ('acuerdo__cliente__referencia', 'Referencia Cliente'),
        ('acuerdo__referencia', 'Referencia Acuerdo'),
        ('acuerdo__monto_total', 'Monto Total Acuerdo'),
        ('acuerdo__numero_cuotas', 'Total Cuotas Acuerdo'),
        ('acuerdo__estado', 'Estado Acuerdo'),
        ('numero_cuota', 'Número de Cuota'),
        ('monto', 'Monto Cuota'),
        ('fecha_pago', 'Fecha de Pago'),
        ('fecha_vencimiento', 'Fecha de Vencimiento'),
        ('estado', 'Estado Cuota'),
        ('comprobante', 'Comprobante'),
        ('observaciones', 'Observaciones')
    ]
}

@login_required
def reportes(request):
    # Determinar si es una solicitud POST o GET
    if request.method == 'POST':
        # Obtener parámetros de filtro de POST
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        tipo_reporte = request.POST.get('tipo_reporte', 'clientes')
        accion = request.POST.get('accion', '')
        campos_seleccionados = request.POST.getlist('campos')
    else:
        # Obtener parámetros de filtro de GET
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        tipo_reporte = request.GET.get('tipo_reporte', 'clientes')
        accion = ''
        campos_seleccionados = request.GET.getlist('campos')
    
    # Obtener campos seleccionados o usar todos por defecto
    if not campos_seleccionados and tipo_reporte in CAMPOS_DISPONIBLES:
        campos_seleccionados = [campo[0] for campo in CAMPOS_DISPONIBLES[tipo_reporte]]
    
    # Establecer fechas por defecto (últimos 30 días)
    if not fecha_fin:
        fecha_fin = timezone.now().strftime('%Y-%m-%d')
    if not fecha_inicio:
        fecha_inicio = (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        
    print(f"Tipo de reporte: {tipo_reporte}, Fecha inicio: {fecha_inicio}, Fecha fin: {fecha_fin}, Acción: {accion}")
    print(f"Campos seleccionados: {campos_seleccionados}")
    
    # Inicializar contexto
    context = {
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'tipo_reporte': tipo_reporte,
        'seccion_activa': 'reportes',
        'CAMPOS_DISPONIBLES': CAMPOS_DISPONIBLES.get(tipo_reporte, []),
        'campos_seleccionados': campos_seleccionados,
    }
    
    # Obtener datos según el tipo de reporte
    if tipo_reporte == 'usuarios':
        wb = Workbook()
        ws = wb.active
        from openpyxl.styles import Font, Alignment, PatternFill
        header_fill = PatternFill(start_color='E5E5E5', end_color='E5E5E5', fill_type='solid')
        header_font = Font(bold=True, color='000000')
        alignment = Alignment(horizontal='center', vertical='center')
        queryset = User.objects.all()
        # Escribir encabezados
        for col_num, campo in enumerate(campos_seleccionados, 1):
            etiqueta = next((etq for c, etq in CAMPOS_DISPONIBLES[tipo_reporte] if c == campo), campo.replace('_', ' ').title())
            cell = ws.cell(row=1, column=col_num, value=etiqueta)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
        # Escribir datos
        for row_num, obj in enumerate(queryset, 2):
            for col_num, campo in enumerate(campos_seleccionados, 1):
                valor = ''
                if campo == 'groups':
                    valor = ', '.join([g.name for g in obj.groups.all()])
                else:
                    valor = getattr(obj, campo, '')
                # Manejar fechas
                if hasattr(valor, 'strftime'):
                    valor = valor.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(valor, bool):
                    valor = 'Sí' if valor else 'No'
                ws.cell(row=row_num, column=col_num, value=valor)

    elif tipo_reporte == 'clientes':
        # Lógica para reporte de clientes
        clientes = Cliente.objects.all()
        
        # Aplicar filtros
        if fecha_inicio and fecha_fin:
            # Convertir las fechas de string a objetos date para asegurar un filtrado correcto
            from datetime import datetime
            fecha_ini = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            # Añadir un día a la fecha final para incluir todo el día en el filtro
            fecha_fin_obj = fecha_fin_obj + timedelta(days=1)
            
            # Filtrar por fecha de registro
            clientes = clientes.filter(
                fecha_registro__range=[fecha_ini, fecha_fin_obj]
            )
            
            # Log de filtrado solo en desarrollo
            if settings.DEBUG:
                logger = logging.getLogger(__name__)
                logger.info(f"Filtrando clientes por fecha: {fecha_ini} a {fecha_fin_obj}")
                logger.info(f"Total de clientes encontrados: {clientes.count()}")
            
        context['datos'] = clientes[:100]  # Limitar a 100 registros para la vista previa
        context['total_registros'] = clientes.count()
        
    elif tipo_reporte == 'gestiones':
        # Lógica para reporte de gestiones
        gestiones = Gestion.objects.select_related('cliente').all()
        
        if fecha_inicio and fecha_fin:
            # Convertir las fechas de string a objetos date
            from datetime import datetime
            fecha_ini = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            # Añadir un día a la fecha final para incluir todo el día en el filtro
            fecha_fin_obj = fecha_fin_obj + timedelta(days=1)
            
            gestiones = gestiones.filter(
                fecha_hora_gestion__range=[fecha_ini, fecha_fin_obj]
            )
            
            # Log de filtrado de gestiones solo en desarrollo
            if settings.DEBUG:
                logger = logging.getLogger(__name__)
                logger.info(f"Filtrando gestiones por fecha: {fecha_ini} a {fecha_fin_obj}")
            # Log de gestiones solo en desarrollo
            if settings.DEBUG:
                logger = logging.getLogger(__name__)
                logger.info(f"Total de gestiones encontradas: {gestiones.count()}")
            
        context['datos'] = gestiones[:100]
        context['total_registros'] = gestiones.count()
        
    elif tipo_reporte == 'acuerdos':
        # Lógica para reporte de acuerdos
        acuerdos = AcuerdoPago.objects.select_related('cliente').all()
        
        if fecha_inicio and fecha_fin:
            # Convertir las fechas de string a objetos date
            from datetime import datetime
            fecha_ini = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            # Añadir un día a la fecha final para incluir todo el día en el filtro
            fecha_fin_obj = fecha_fin_obj + timedelta(days=1)
            
            acuerdos = acuerdos.filter(
                fecha_creacion__range=[fecha_ini, fecha_fin_obj]
            )
            
            # Log de filtrado de acuerdos solo en desarrollo
            if settings.DEBUG:
                logger = logging.getLogger(__name__)
                logger.info(f"Filtrando acuerdos por fecha: {fecha_ini} a {fecha_fin_obj}")
            # Log de acuerdos solo en desarrollo
            if settings.DEBUG:
                logger = logging.getLogger(__name__)
                logger.info(f"Total de acuerdos encontrados: {acuerdos.count()}")
            
        context['datos'] = acuerdos[:100]
        context['total_registros'] = acuerdos.count()
        
    elif tipo_reporte == 'pagos':
        # Lógica para reporte de pagos
        pagos = CuotaAcuerdo.objects.select_related('acuerdo__cliente').filter(
            estado='pagada'
        )
        
        if fecha_inicio and fecha_fin:
            # Convertir las fechas de string a objetos date
            from datetime import datetime
            fecha_ini = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            # Añadir un día a la fecha final para incluir todo el día en el filtro
            fecha_fin_obj = fecha_fin_obj + timedelta(days=1)
            
            # Usar range sin el lookup date__ ya que fecha_pago es DateField
            pagos = pagos.filter(
                fecha_pago__range=[fecha_ini, fecha_fin_obj]
            )
            
            # Log de filtrado de pagos solo en desarrollo
            if settings.DEBUG:
                logger = logging.getLogger(__name__)
                logger.info(f"Filtrando pagos por fecha: {fecha_ini} a {fecha_fin_obj}")
            # Log de pagos solo en desarrollo
            if settings.DEBUG:
                logger = logging.getLogger(__name__)
                logger.info(f"Total de pagos encontrados: {pagos.count()}")
            
        context['datos'] = pagos[:100]
        context['total_registros'] = pagos.count()
    
    return render(request, 'core/reportes/reportes.html', context)

@login_required
def exportar_excel(request):
    if request.method == 'POST':
        tipo_reporte = request.POST.get('tipo_reporte', 'clientes')
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        campos_seleccionados = request.POST.getlist('campos')
    else:
        # Si se accede por GET, obtener los parámetros de la URL
        tipo_reporte = request.GET.get('tipo_reporte', 'clientes')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        campos_seleccionados = request.GET.getlist('campos')
        
        # Si no hay parámetros, redirigir a la página de reportes
        if not (tipo_reporte and (fecha_inicio or fecha_fin)):
            return redirect('core:reportes')
    
    print(f"Exportando Excel - Tipo: {tipo_reporte}, Fecha inicio: {fecha_inicio}, Fecha fin: {fecha_fin}")
    print(f"Campos seleccionados: {campos_seleccionados}")
    
    # Crear el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = tipo_reporte.capitalize()
    
    # Estilos para el encabezado
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Obtener etiquetas de los campos seleccionados
    if not campos_seleccionados and tipo_reporte in CAMPOS_DISPONIBLES:
        campos_seleccionados = [campo[0] for campo in CAMPOS_DISPONIBLES[tipo_reporte]]
    
    # Mapeo de campos a funciones de obtención de valor
    def get_valor(objeto, campo, valor_default=''):
        try:
            # Manejar campos con relaciones (usando __ como separador)
            if '__' in campo:
                partes = campo.split('__')
                obj = objeto
                
                for i, parte in enumerate(partes):
                    if obj is None:
                        return valor_default
                    
                    # Obtener el atributo del objeto actual
                    obj = getattr(obj, parte, None)
                    
                    # Si es None, retornar valor por defecto
                    if obj is None:
                        return valor_default
                    
                    # Si es un queryset, obtener el primer elemento
                    if hasattr(obj, 'all') and not isinstance(obj, (str, int, float, bool, dict, list)):
                        if hasattr(obj, 'first'):
                            obj = obj.first()
                            if obj is None:
                                return valor_default
                
                # Si llegamos aquí, devolver el valor final
                if hasattr(obj, '__str__'):
                    return str(obj)
                return obj if obj is not None else valor_default
            
            # Manejar campos directos
            valor = getattr(objeto, campo, valor_default)
            
            # Si es una fecha/hora, formatear sin microsegundos
            if hasattr(valor, 'strftime'):
                return valor.strftime('%Y-%m-%d %H:%M:%S')
            # Si es un objeto, obtener su representación como string
            elif hasattr(valor, '__str__') and not isinstance(valor, (str, int, float, bool)):
                return str(valor)
                
            return valor if valor is not None else valor_default
                
        except Exception as e:
            print(f"Error al obtener valor para {campo}: {str(e)}")
            return valor_default
    
    # Obtener datos según el tipo de reporte
    if tipo_reporte == 'clientes':
        # Obtener datos
        queryset = Cliente.objects.all()
        if fecha_inicio and fecha_fin:
            # Convertir las fechas de string a objetos date
            from datetime import datetime
            fecha_ini = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            # Añadir un día a la fecha final para incluir todo el día en el filtro
            fecha_fin_obj = fecha_fin_obj + timedelta(days=1)
            
            # Usar range en lugar de date__range para asegurar un filtrado correcto
            queryset = queryset.filter(
                fecha_registro__range=[fecha_ini, fecha_fin_obj]
            )
            
            # Log de filtrado solo en desarrollo
            if settings.DEBUG:
                logger = logging.getLogger(__name__)
                logger.info(f"Filtrando clientes por fecha: {fecha_ini} a {fecha_fin_obj}")
                logger.info(f"Total de clientes encontrados: {queryset.count()}")
        
        # Escribir encabezados
        for col_num, campo in enumerate(campos_seleccionados, 1):
            # Buscar la etiqueta del campo
            etiqueta = next((etq for c, etq in CAMPOS_DISPONIBLES[tipo_reporte] if c == campo), campo.replace('_', ' ').title())
            cell = ws.cell(row=1, column=col_num, value=etiqueta)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
        
        # Escribir datos
        for row_num, obj in enumerate(queryset, 2):
            for col_num, campo in enumerate(campos_seleccionados, 1):
                valor = ''
                if campo == 'ultima_gestion':
                    ultima_gestion = obj.gestiones.order_by('-fecha_hora_gestion').first()
                    if ultima_gestion:
                        valor = f"{ultima_gestion.fecha_hora_gestion.strftime('%Y-%m-%d %H:%M')} - {ultima_gestion.usuario_gestion.username if ultima_gestion.usuario_gestion else ''} - {ultima_gestion.tipo_gestion_n1} - {ultima_gestion.estado_contacto}"
                    else:
                        valor = ''
                elif campo == 'cliente__documento':
                    valor = obj.cliente.documento if obj.cliente else ''
                elif campo == 'cliente__nombre_completo':
                    valor = obj.cliente.nombre_completo if obj.cliente else ''
                elif campo == 'cliente__ciudad':
                    valor = obj.cliente.ciudad if obj.cliente else ''
                elif campo == 'cliente__email':
                    valor = obj.cliente.email if obj.cliente else ''
                elif campo == 'cliente__telefono_celular':
                    valor = obj.cliente.telefono_celular if obj.cliente else ''
                elif campo == 'cliente__referencia':
                    valor = obj.cliente.referencia if obj.cliente else ''
                elif campo == 'cliente__direccion_1':
                    valor = obj.cliente.direccion_1 if obj.cliente else ''
                elif campo == 'cliente__observaciones_adicionales':
                    valor = obj.cliente.observaciones_adicionales if obj.cliente else ''
                elif campo == 'cuotas_pagadas':
                    valor = obj.cuotas.filter(estado='pagada').count()
                elif campo == 'monto_pagado':
                    valor = sum([c.monto for c in obj.cuotas.filter(estado='pagada')])
                elif campo == 'fecha_ultimo_pago':
                    ultima = obj.cuotas.filter(estado='pagada', fecha_pago__isnull=False).order_by('-fecha_pago').first()
                    valor = ultima.fecha_pago.strftime('%Y-%m-%d') if ultima and ultima.fecha_pago else ''
                elif campo == 'estado_cuotas':
                    total = obj.cuotas.count()
                    pagadas = obj.cuotas.filter(estado='pagada').count()
                    vencidas = obj.cuotas.filter(estado='vencida').count()
                    if total == pagadas:
                        valor = 'Pagado'
                    elif vencidas > 0:
                        valor = f'{vencidas} vencida(s)'
                    else:
                        valor = f'{pagadas} pagada(s) de {total}'
                elif campo == 'comprobantes_pago':
                    comprobantes = obj.cuotas.filter(comprobante_pago__isnull=False).exclude(comprobante_pago='')
                    valor = ', '.join([c.comprobante_pago.url for c in comprobantes if c.comprobante_pago])
                elif campo == 'fecha_ultima_gestion':
                    ultima_gestion = Gestion.objects.filter(cliente=obj.cliente).order_by('-fecha_hora_gestion').first()
                    valor = ultima_gestion.fecha_hora_gestion.strftime('%Y-%m-%d %H:%M') if ultima_gestion and ultima_gestion.fecha_hora_gestion else ''
                elif campo == 'usuario_ultima_gestion':
                    ultima_gestion = Gestion.objects.filter(cliente=obj.cliente).order_by('-fecha_hora_gestion').first()
                    valor = ultima_gestion.usuario_gestion.username if ultima_gestion and ultima_gestion.usuario_gestion else ''
                elif campo == 'observaciones_ultima_gestion':
                    ultima_gestion = Gestion.objects.filter(cliente=obj.cliente).order_by('-fecha_hora_gestion').first()
                    valor = ultima_gestion.observaciones if ultima_gestion and ultima_gestion.observaciones else ''
                else:
                    valor = get_valor(obj, campo)
                ws.cell(row=row_num, column=col_num, value=valor)
    
    elif tipo_reporte == 'gestiones':
        # Precargar relaciones para mejorar el rendimiento
        # Primero aplicar los filtros
        queryset = Gestion.objects.select_related('cliente', 'usuario_gestion').prefetch_related('acuerdos')
        
        if fecha_inicio and fecha_fin:
            # Convertir las fechas de string a objetos date
            from datetime import datetime
            fecha_ini = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            # Añadir un día a la fecha final para incluir todo el día en el filtro
            fecha_fin_obj = fecha_fin_obj + timedelta(days=1)
            
            queryset = queryset.filter(
                fecha_hora_gestion__range=[fecha_ini, fecha_fin_obj]
            )
        
        # Forzar la carga de los acuerdos para evitar consultas adicionales
        gestiones_con_acuerdos = []
        for gestion in queryset:
            gestion.acuerdos_list = list(gestion.acuerdos.all())
            if gestion.acuerdos_list:
                gestion.acuerdo = gestion.acuerdos_list[0]
            else:
                gestion.acuerdo = None
            gestiones_con_acuerdos.append(gestion)
            
        queryset = gestiones_con_acuerdos
        
        # Log de gestiones solo en desarrollo
        if settings.DEBUG:
            logger = logging.getLogger(__name__)
            logger.info(f"Total de gestiones encontradas: {len(queryset)}")
        
        # Escribir encabezados
        for col_num, campo in enumerate(campos_seleccionados, 1):
            etiqueta = next((etq for c, etq in CAMPOS_DISPONIBLES[tipo_reporte] if c == campo), campo.replace('_', ' ').title())
            cell = ws.cell(row=1, column=col_num, value=etiqueta)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
        
        # Escribir datos
        for row_num, obj in enumerate(queryset, 2):
            # Precargar manualmente la relación de acuerdos si es necesario
            if 'acuerdos__' in ' '.join(campos_seleccionados):
                if not hasattr(obj, '_acuerdos_cargados'):
                    obj.acuerdos_list = list(obj.acuerdos.all())
                    obj._acuerdos_cargados = True
                # Usar el primer acuerdo si existe
                if hasattr(obj, 'acuerdos_list') and obj.acuerdos_list:
                    obj.acuerdo = obj.acuerdos_list[0]
                else:
                    obj.acuerdo = None
            
            for col_num, campo in enumerate(campos_seleccionados, 1):
                valor = get_valor(obj, campo)

                # Manejar campos especiales
                if campo == 'comprobante_pago' and getattr(obj, 'comprobante_pago', None):
                    # Si hay archivo, poner la URL relativa para descarga
                    valor = obj.comprobante_pago.url if obj.comprobante_pago else ''
                elif campo == 'cliente__referencia':
                    # Mostrar todas las referencias asociadas al documento del cliente
                    if obj.cliente:
                        referencias = Cliente.objects.filter(documento=obj.cliente.documento).exclude(referencia__isnull=True).exclude(referencia='').values_list('referencia', flat=True).distinct()
                        valor = ', '.join(referencias)
                    else:
                        valor = ''
                elif campo == 'cliente__observaciones_adicionales':
                    valor = obj.cliente.observaciones_adicionales if obj.cliente and hasattr(obj.cliente, 'observaciones_adicionales') else ''
                elif campo == 'cliente__direccion_1':
                    valor = obj.cliente.direccion_1 if obj.cliente and hasattr(obj.cliente, 'direccion_1') else ''
                elif campo == 'cliente__ciudad':
                    valor = obj.cliente.ciudad if obj.cliente and hasattr(obj.cliente, 'ciudad') else ''

                # Manejar valores nulos
                if valor is None:
                    valor = ''
                # Manejar fechas
                elif hasattr(valor, 'strftime'):
                    valor = valor.strftime('%Y-%m-%d %H:%M:%S')
                # Manejar booleanos
                elif isinstance(valor, bool):
                    valor = 'Sí' if valor else 'No'
                # Para relaciones, asegurarse de obtener el string
                elif hasattr(valor, '__str__') and not isinstance(valor, (str, int, float, bool)):
                    valor = str(valor)

                print(f"Fila {row_num}, Col {col_num} ({campo}): {valor}")
                ws.cell(row=row_num, column=col_num, value=valor)
    
    elif tipo_reporte == 'acuerdos':
        queryset = AcuerdoPago.objects.select_related('cliente').all()
        if fecha_inicio and fecha_fin:
            # Convertir las fechas de string a objetos date
            from datetime import datetime
            fecha_ini = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            # Añadir un día a la fecha final para incluir todo el día en el filtro
            fecha_fin_obj = fecha_fin_obj + timedelta(days=1)
            
            queryset = queryset.filter(
                fecha_creacion__range=[fecha_ini, fecha_fin_obj]
            )
        
        # Escribir encabezados
        for col_num, campo in enumerate(campos_seleccionados, 1):
            etiqueta = next((etq for c, etq in CAMPOS_DISPONIBLES[tipo_reporte] if c == campo), campo.replace('_', ' ').title())
            cell = ws.cell(row=1, column=col_num, value=etiqueta)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
        
        # Escribir datos
        for row_num, obj in enumerate(queryset, 2):
            for col_num, campo in enumerate(campos_seleccionados, 1):
                valor = get_valor(obj, campo)
                ws.cell(row=row_num, column=col_num, value=valor)
    
    elif tipo_reporte == 'pagos':
        queryset = CuotaAcuerdo.objects.select_related('acuerdo__cliente').all()
        if fecha_inicio and fecha_fin:
            # Convertir las fechas de string a objetos date
            from datetime import datetime
            fecha_ini = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            # Añadir un día a la fecha final para incluir todo el día en el filtro
            fecha_fin_obj = fecha_fin_obj + timedelta(days=1)
            
            # Usar range sin el lookup date__ ya que fecha_pago es DateField
            queryset = queryset.filter(
                fecha_pago__range=[fecha_ini, fecha_fin_obj]
            )
        
        # Escribir encabezados
        for col_num, campo in enumerate(campos_seleccionados, 1):
            etiqueta = next((etq for c, etq in CAMPOS_DISPONIBLES[tipo_reporte] if c == campo), campo.replace('_', ' ').title())
            cell = ws.cell(row=1, column=col_num, value=etiqueta)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
        
        # Escribir datos
        for row_num, obj in enumerate(queryset, 2):
            for col_num, campo in enumerate(campos_seleccionados, 1):
                valor = get_valor(obj, campo)
                ws.cell(row=row_num, column=col_num, value=valor)
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width if adjusted_width < 40 else 40
    
    # Crear respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=reporte_{tipo_reporte}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response
