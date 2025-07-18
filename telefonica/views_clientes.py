from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils import timezone
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import json

from .models import ClientesUpgrade, ClientesPrePos
from .forms import ClientesUpgradeForm, ClientesPrePosForm

# Función auxiliar para verificar permisos
def tiene_permiso_clientes(user):
    return user.groups.filter(name__in=['asesor', 'backoffice', 'administrador']).exists() or user.is_superuser

@login_required
@user_passes_test(tiene_permiso_clientes)
def clientes_lista(request):
    """
    Vista principal para la página de clientes que muestra tanto ClientesUpgrade como ClientesPrePos
    """
    # Obtener parámetros de búsqueda
    tipo_cliente = request.GET.get('tipo_cliente', 'upgrade')
    query = request.GET.get('q', '')
    
    # Inicializar queryset según el tipo de cliente seleccionado
    if tipo_cliente == 'prepos':
        clientes_list = ClientesPrePos.objects.all().order_by('-fecha_creacion')
        # Aplicar filtros de búsqueda para ClientesPrePos
        if query:
            clientes_list = clientes_list.filter(telefono__icontains=query)
    else:  # Por defecto, mostrar clientes upgrade
        clientes_list = ClientesUpgrade.objects.all().order_by('-fecha_creacion')
        # Aplicar filtros de búsqueda para ClientesUpgrade
        if query:
            clientes_list = clientes_list.filter(
                Q(nombre_cliente__icontains=query) | 
                Q(documento__icontains=query) | 
                Q(id_base__icontains=query) |
                Q(tel_contacto_1__icontains=query) |
                Q(celular_contacto__icontains=query)
            )
    
    # Paginación
    paginator = Paginator(clientes_list, 20)  # 20 clientes por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Preparar contexto
    context = {
        'page_obj': page_obj,
        'tipo_cliente': tipo_cliente,
        'query': query,
        'total_clientes': clientes_list.count(),
    }
    
    return render(request, 'telefonica/clientes/clientes_lista.html', context)

@login_required
@user_passes_test(tiene_permiso_clientes)
def cliente_upgrade_crear(request):
    """
    Vista para crear un nuevo cliente upgrade
    """
    if request.method == 'POST':
        form = ClientesUpgradeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cliente Upgrade creado exitosamente.')
            return redirect('telefonica:clientes_lista')
    else:
        form = ClientesUpgradeForm()
    
    return render(request, 'telefonica/clientes/cliente_upgrade_form.html', {
        'form': form,
        'titulo': 'Crear Cliente Upgrade',
        'accion': 'Crear'
    })

@login_required
@user_passes_test(tiene_permiso_clientes)
def cliente_upgrade_editar(request, pk):
    """
    Vista para editar un cliente upgrade existente
    """
    cliente = get_object_or_404(ClientesUpgrade, pk=pk)
    
    if request.method == 'POST':
        form = ClientesUpgradeForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cliente Upgrade actualizado exitosamente.')
            return redirect('telefonica:clientes_lista')
    else:
        form = ClientesUpgradeForm(instance=cliente)
    
    return render(request, 'telefonica/clientes/cliente_upgrade_form.html', {
        'form': form,
        'cliente': cliente,
        'titulo': 'Editar Cliente Upgrade',
        'accion': 'Actualizar'
    })

@login_required
@user_passes_test(tiene_permiso_clientes)
def cliente_upgrade_eliminar(request, pk):
    """
    Vista para eliminar un cliente upgrade
    """
    cliente = get_object_or_404(ClientesUpgrade, pk=pk)
    
    if request.method == 'POST':
        cliente.delete()
        messages.success(request, 'Cliente Upgrade eliminado exitosamente.')
        return redirect('telefonica:clientes_lista')
    
    return render(request, 'telefonica/clientes/cliente_confirmar_eliminar.html', {
        'cliente': cliente,
        'tipo_cliente': 'upgrade'
    })

@login_required
@user_passes_test(tiene_permiso_clientes)
def cliente_prepos_crear(request):
    """
    Vista para crear un nuevo cliente prepos
    """
    if request.method == 'POST':
        form = ClientesPrePosForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cliente PrePos creado exitosamente.')
            return redirect('telefonica:clientes_lista')
    else:
        form = ClientesPrePosForm()
    
    return render(request, 'telefonica/clientes/cliente_prepos_form.html', {
        'form': form,
        'titulo': 'Crear Cliente PrePos',
        'accion': 'Crear'
    })

@login_required
@user_passes_test(tiene_permiso_clientes)
def cliente_prepos_editar(request, pk):
    """
    Vista para editar un cliente prepos existente
    """
    cliente = get_object_or_404(ClientesPrePos, pk=pk)
    
    if request.method == 'POST':
        form = ClientesPrePosForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cliente PrePos actualizado exitosamente.')
            return redirect('telefonica:clientes_lista')
    else:
        form = ClientesPrePosForm(instance=cliente)
    
    return render(request, 'telefonica/clientes/cliente_prepos_form.html', {
        'form': form,
        'cliente': cliente,
        'titulo': 'Editar Cliente PrePos',
        'accion': 'Actualizar'
    })

@login_required
@user_passes_test(tiene_permiso_clientes)
def cliente_prepos_eliminar(request, pk):
    """
    Vista para eliminar un cliente prepos
    """
    cliente = get_object_or_404(ClientesPrePos, pk=pk)
    
    if request.method == 'POST':
        cliente.delete()
        messages.success(request, 'Cliente PrePos eliminado exitosamente.')
        return redirect('telefonica:clientes_lista')
    
    return render(request, 'telefonica/clientes/cliente_confirmar_eliminar.html', {
        'cliente': cliente,
        'tipo_cliente': 'prepos'
    })

@login_required
@user_passes_test(tiene_permiso_clientes)
def carga_clientes_upgrade(request):
    """
    Vista para cargar clientes upgrade masivamente desde un archivo Excel
    """
    context = {
        'titulo': 'Carga masiva de clientes Upgrade',
    }
    
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        extension = os.path.splitext(archivo.name)[1].lower()
        
        if extension not in ['.xlsx', '.csv']:
            messages.error(request, 'El archivo debe ser Excel (.xlsx) o CSV (.csv)')
            return render(request, 'telefonica/clientes/carga_clientes.html', context)
        
        try:
            # Leer el archivo según su extensión
            if extension == '.xlsx':
                df = pd.read_excel(archivo)
            else:  # CSV
                df = pd.read_csv(archivo)
            
            # Procesar datos
            nuevos = 0
            actualizados = 0
            errores = []
            
            # Renombrar columnas a minúsculas y quitar espacios/acentos
            df.columns = [c.strip().lower().replace(' ', '_') for c in df.columns]
            
            # Verificar columnas requeridas
            columnas_requeridas = ['id_base', 'nro_registro', 'nombre_cliente', 'documento']
            for columna in columnas_requeridas:
                if columna not in df.columns:
                    messages.error(request, f'El archivo no contiene la columna requerida: {columna}')
                    return render(request, 'telefonica/clientes/carga_clientes.html', context)
            
            # Procesar cada registro
            for index, row in df.iterrows():
                try:
                    # Datos básicos requeridos
                    id_base = str(row.get('id_base', '')).strip()
                    nro_registro = str(row.get('nro_registro', '')).strip()
                    nombre_cliente = str(row.get('nombre_cliente', '')).strip()
                    documento = str(row.get('documento', '')).strip()
                    
                    if not id_base or not nro_registro or not nombre_cliente or not documento:
                        errores.append(f'Fila {index + 2}: ID Base, Número de Registro, Nombre Cliente y Documento son obligatorios')
                        continue
                    
                    # Preparar datos para crear/actualizar
                    datos = {}
                    for campo in ClientesUpgradeForm.Meta.fields:
                        if campo in row and not pd.isna(row[campo]):
                            datos[campo] = row[campo]
                    
                    # Intentar actualizar o crear usando nro_registro como identificador único
                    cliente, created = ClientesUpgrade.objects.update_or_create(
                        nro_registro=nro_registro,
                        defaults=datos
                    )
                    
                    if created:
                        nuevos += 1
                    else:
                        actualizados += 1
                        
                except Exception as e:
                    errores.append(f'Fila {index + 2}: {str(e)}')
            
            # Preparar mensaje de resultado
            if errores:
                messages.warning(request, f'Carga completada con errores. {nuevos} nuevos, {actualizados} actualizados. {len(errores)} errores.')
                context['errores'] = errores
            else:
                messages.success(request, f'Carga exitosa: {nuevos} nuevos, {actualizados} actualizados.')
            
            context['resumen'] = {
                'nuevos': nuevos,
                'actualizados': actualizados,
                'errores': len(errores)
            }
            
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
    
    return render(request, 'telefonica/clientes/carga_clientes.html', context)

@login_required
@user_passes_test(tiene_permiso_clientes)
def carga_clientes_prepos(request):
    """
    Vista para cargar clientes prepos masivamente desde un archivo Excel
    """
    context = {
        'titulo': 'Carga masiva de clientes PrePos',
    }
    
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        extension = os.path.splitext(archivo.name)[1].lower()
        
        if extension not in ['.xlsx', '.csv']:
            messages.error(request, 'El archivo debe ser Excel (.xlsx) o CSV (.csv)')
            return render(request, 'telefonica/clientes/carga_clientes_prepos.html', context)
        
        try:
            # Leer el archivo según su extensión
            if extension == '.xlsx':
                df = pd.read_excel(archivo)
            else:  # CSV
                df = pd.read_csv(archivo)
            
            # Procesar datos
            nuevos = 0
            errores = []
            duplicados = 0
            
            # Renombrar columnas a minúsculas y quitar espacios/acentos
            df.columns = [c.strip().lower().replace(' ', '_') for c in df.columns]
            
            # Verificar columna requerida
            if 'telefono' not in df.columns:
                messages.error(request, 'El archivo no contiene la columna requerida: telefono')
                return render(request, 'telefonica/clientes/carga_clientes_prepos.html', context)
            
            # Procesar cada registro
            for index, row in df.iterrows():
                try:
                    telefono = str(row.get('telefono', '')).strip()
                    
                    if not telefono:
                        errores.append(f'Fila {index + 2}: Teléfono es obligatorio')
                        continue
                    
                    # Verificar si ya existe
                    if ClientesPrePos.objects.filter(telefono=telefono).exists():
                        duplicados += 1
                        continue
                    
                    # Crear nuevo cliente
                    ClientesPrePos.objects.create(telefono=telefono)
                    nuevos += 1
                        
                except Exception as e:
                    errores.append(f'Fila {index + 2}: {str(e)}')
            
            # Preparar mensaje de resultado
            if errores:
                messages.warning(request, f'Carga completada con errores. {nuevos} nuevos, {duplicados} duplicados. {len(errores)} errores.')
                context['errores'] = errores
            else:
                messages.success(request, f'Carga exitosa: {nuevos} nuevos, {duplicados} duplicados ignorados.')
            
            context['resumen'] = {
                'nuevos': nuevos,
                'duplicados': duplicados,
                'errores': len(errores)
            }
            
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
    
    return render(request, 'telefonica/clientes/carga_clientes_prepos.html', context)

@login_required
@user_passes_test(tiene_permiso_clientes)
def descargar_plantilla_upgrade(request):
    """
    Vista para descargar la plantilla Excel para carga de clientes upgrade
    """
    # Crear un libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Clientes Upgrade"
    
    # Definir encabezados basados en los campos del modelo
    encabezados = [
        'id_base', 'nro_registro', 'campana', 'grupo_campana', 'estrategia', 'nombre_cliente',
        'tipo_documento', 'documento', 'direccion', 'estrato', 'barrio', 'departamento',
        'ciudad', 'producto', 'puertos_disponibles', 'promedio_fact', 'mx_tenencia_cuenta',
        'tel_contacto_1', 'tel_contacto_2', 'tel_contacto_3', 'celular_contacto'
    ]
    
    # Aplicar estilos al encabezado
    for col_num, encabezado in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col_num, value=encabezado)
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal='center')
        celda.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        # Ajustar ancho de columna
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    # Crear respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=plantilla_clientes_upgrade.xlsx'
    wb.save(response)
    
    return response

@login_required
@user_passes_test(tiene_permiso_clientes)
def descargar_plantilla_prepos(request):
    """
    Vista para descargar la plantilla Excel para carga de clientes prepos
    """
    # Crear un libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Clientes PrePos"
    
    # Definir encabezados
    encabezados = ['telefono']
    
    # Aplicar estilos al encabezado
    for col_num, encabezado in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col_num, value=encabezado)
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal='center')
        celda.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        # Ajustar ancho de columna
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    # Crear respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=plantilla_clientes_prepos.xlsx'
    wb.save(response)
    
    return response


@login_required
def buscar_cliente_upgrade_por_documento(request):
    """
    Endpoint para buscar un cliente por número de documento o número de registro
    y devolver sus datos en formato JSON para autocompletar formularios.
    """
    documento = request.GET.get('documento', '')
    nro_registro = request.GET.get('nro_registro', '')
    
    if nro_registro:
        # Buscar por número de registro si está disponible
        try:
            cliente = ClientesUpgrade.objects.get(nro_registro=nro_registro)
            encontrado = True
        except ClientesUpgrade.DoesNotExist:
            encontrado = False
            return JsonResponse({
                'encontrado': encontrado,
                'mensaje': 'Cliente no encontrado en la base de datos.'
            })
    elif documento:
        # Buscar por documento si no hay número de registro
        try:
            cliente = ClientesUpgrade.objects.get(documento=documento)
            encontrado = True
        except ClientesUpgrade.DoesNotExist:
            encontrado = False
            return JsonResponse({
                'encontrado': encontrado,
                'mensaje': 'Cliente no encontrado en la base de datos.'
            })
    else:
        # Si no se proporciona ningún parámetro de búsqueda
        return JsonResponse({
            'encontrado': False,
            'mensaje': 'No se proporcionó un documento o número de registro válido.'
        })
    
    # Si se encontró el cliente, devolver sus datos
    if encontrado:
        return JsonResponse({
            'encontrado': True,
            'cliente': {
                'id': cliente.id,
                'id_base': cliente.id_base,
                'nro_registro': cliente.nro_registro,
                'nombre_cliente': cliente.nombre_cliente,
                'tipo_documento': cliente.tipo_documento,
                'documento': cliente.documento,
                'direccion': cliente.direccion or '',
                'estrato': str(cliente.estrato) if cliente.estrato else '',
                'barrio': cliente.barrio or '',
                'departamento': cliente.departamento or '',
                'ciudad': cliente.ciudad or '',
                'producto': cliente.producto or '',
                'tel_contacto_1': cliente.tel_contacto_1 or '',
                'celular_contacto': cliente.celular_contacto or ''
            }
        })


@login_required
def buscar_cliente_prepos_por_telefono(request):
    """
    Endpoint para buscar un cliente PrePos por número de teléfono
    y devolver sus datos en formato JSON para autocompletar formularios.
    """
    telefono = request.GET.get('telefono', '')
    
    if telefono:
        # Buscar por teléfono
        try:
            cliente = ClientesPrePos.objects.get(telefono=telefono)
            encontrado = True
        except ClientesPrePos.DoesNotExist:
            encontrado = False
            return JsonResponse({
                'encontrado': encontrado,
                'mensaje': 'Cliente no encontrado en la base de datos.'
            })
    else:
        # Si no se proporciona ningún parámetro de búsqueda
        return JsonResponse({
            'encontrado': False,
            'mensaje': 'No se proporcionó un número de teléfono válido.'
        })
    
    # Si se encontró el cliente, devolver sus datos
    if encontrado:
        return JsonResponse({
            'encontrado': True,
            'cliente': {
                'id': cliente.id,
                'telefono': cliente.telefono,
                'estado': 'dentro_base'
            }
        })