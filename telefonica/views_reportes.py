from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import pandas as pd
import logging

from .models import (
    VentaPortabilidad, VentaPrePos, VentaUpgrade, 
    Agendamiento, GestionAgendamiento, GestionAsesor, 
    GestionBackoffice, Escalamiento, Comision,
    Planes_portabilidad, ClientesPrePos, ClientesUpgrade
)
from django.contrib.auth.models import User, Group

# Diccionario con los campos disponibles para cada tipo de reporte de telefónica
CAMPOS_DISPONIBLES = {
    'ventas_portabilidad': [
        ('numero', 'Número'),
        ('tipo_documento', 'Tipo Documento'),
        ('documento', 'Documento'),
        ('fecha_expedicion', 'Fecha Expedición'),
        ('nombre_completo', 'Nombre Completo'),
        ('telefono_legalizacion', 'Teléfono Legalización'),
        ('numero_a_portar', 'Número a Portar'),
        ('nip', 'NIP'),
        ('fecha_entrega', 'Fecha Entrega'),
        ('fecha_ventana_cambio', 'Fecha Ventana Cambio'),
        ('numero_orden', 'Número Orden'),
        ('base_origen', 'Base Origen'),
        ('usuario_greta', 'Usuario Greta'),
        ('plan_nombre', 'Plan Nombre'),
        ('plan_codigo', 'Plan Código'),
        ('plan_cfm', 'Plan CFM'),
        ('plan_cfm_sin_iva', 'Plan CFM sin IVA'),
        ('estado_venta', 'Estado Venta'),
        ('estado_logistica', 'Estado Logística'),
        ('agente__username', 'Agente'),
        ('backoffice__username', 'Backoffice'),
        ('fecha_creacion', 'Fecha Creación'),
        ('fecha_actualizacion', 'Fecha Actualización'),
        ('observacion', 'Observación'),
    ],
    'ventas_prepos': [
        ('numero', 'Número'),
        ('tipo_documento', 'Tipo Documento'),
        ('documento', 'Documento'),
        ('fecha_expedicion', 'Fecha Expedición'),
        ('nombre_completo', 'Nombre Completo'),
        ('telefono_legalizacion', 'Teléfono Legalización'),
        ('tipo_cliente', 'Tipo Cliente'),
        ('cliente_base__telefono', 'Cliente Base Teléfono'),
        ('numero_orden', 'Número Orden'),
        ('base_origen', 'Base Origen'),
        ('usuario_greta', 'Usuario Greta'),
        ('plan_nombre', 'Plan Nombre'),
        ('plan_codigo', 'Plan Código'),
        ('plan_cfm', 'Plan CFM'),
        ('plan_cfm_sin_iva', 'Plan CFM sin IVA'),
        ('estado_venta', 'Estado Venta'),
        ('agente__username', 'Agente'),
        ('fecha_creacion', 'Fecha Creación'),
        ('fecha_actualizacion', 'Fecha Actualización'),
        ('observacion', 'Observación'),
    ],
    'ventas_upgrade': [
        ('numero', 'Número'),
        ('tipo_documento', 'Tipo Documento'),
        ('documento', 'Documento'),
        ('fecha_expedicion', 'Fecha Expedición'),
        ('nombre_completo', 'Nombre Completo'),
        ('telefono_legalizacion', 'Teléfono Legalización'),
        ('valor_plan_anterior', 'Valor Plan Anterior'),
        ('tipo_cliente', 'Tipo Cliente'),
        ('cliente_base__nombre_cliente', 'Cliente Base Nombre'),
        ('cliente_base__documento', 'Cliente Base Documento'),
        ('numero_orden', 'Número Orden'),
        ('base_origen', 'Base Origen'),
        ('usuario_greta', 'Usuario Greta'),
        ('plan_nombre', 'Plan Nombre'),
        ('plan_codigo', 'Plan Código'),
        ('plan_cfm', 'Plan CFM'),
        ('plan_cfm_sin_iva', 'Plan CFM sin IVA'),
        ('estado_venta', 'Estado Venta'),
        ('agente__username', 'Agente'),
        ('fecha_creacion', 'Fecha Creación'),
        ('fecha_actualizacion', 'Fecha Actualización'),
        ('observacion', 'Observación'),
    ],
    'agendamientos': [
        ('Estado_agendamiento', 'Estado'),
        ('tipo_venta', 'Tipo Venta'),
        ('nombre_cliente', 'Nombre Cliente'),
        ('telefono_contacto', 'Teléfono Contacto'),
        ('fecha_volver_a_llamar', 'Fecha Volver a Llamar'),
        ('hora_volver_a_llamar', 'Hora Volver a Llamar'),
        ('observaciones', 'Observaciones'),
        ('agente__username', 'Agente'),
        ('fecha_creacion', 'Fecha Creación'),
        ('fecha_actualizacion', 'Fecha Actualización'),
    ],
    'comisiones': [
        ('venta__numero', 'Número Venta'),
        ('venta__nombre_completo', 'Cliente'),
        ('venta__documento', 'Documento Cliente'),
        ('agente__username', 'Agente'),
        ('monto', 'Monto'),
        ('estado', 'Estado'),
        ('fecha_creacion', 'Fecha Creación'),
        ('fecha_pago', 'Fecha Pago'),
    ],
    'escalamientos': [
        ('venta__numero', 'Número Venta'),
        ('venta__nombre_completo', 'Cliente'),
        ('venta__documento', 'Documento Cliente'),
        ('tipo_escalamiento', 'Tipo Escalamiento'),
        ('descripcion', 'Descripción'),
        ('fecha_escalamiento', 'Fecha Escalamiento'),
        ('fecha_solucion', 'Fecha Solución'),
        ('solucionado', 'Solucionado'),
    ],
    'planes': [
        ('codigo', 'Código'),
        ('nombre_plan', 'Nombre Plan'),
        ('caracteristicas', 'Características'),
        ('CFM', 'CFM'),
        ('CFM_sin_iva', 'CFM sin IVA'),
        ('tipo_plan', 'Tipo Plan'),
        ('estado', 'Estado'),
        ('fecha_creacion', 'Fecha Creación'),
        ('fecha_actualizacion', 'Fecha Actualización'),
    ],
}

@login_required
def reportes(request):
    """Vista principal para generar reportes de telefónica"""
    # Determinar si es una solicitud POST o GET
    if request.method == 'POST':
        # Obtener parámetros de filtro de POST
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        tipo_reporte = request.POST.get('tipo_reporte', 'ventas_portabilidad')
        accion = request.POST.get('accion', '')
        campos_seleccionados = request.POST.getlist('campos')
    else:
        # Obtener parámetros de filtro de GET
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        tipo_reporte = request.GET.get('tipo_reporte', 'ventas_portabilidad')
        accion = ''
        campos_seleccionados = request.GET.getlist('campos')
    
    # Establecer fechas por defecto si no se proporcionan
    if not fecha_inicio:
        fecha_inicio = (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not fecha_fin:
        fecha_fin = timezone.now().strftime('%Y-%m-%d')
    
    # Si no hay campos seleccionados, seleccionar todos por defecto
    campos_disponibles_tipo = CAMPOS_DISPONIBLES.get(tipo_reporte, [])
    if not campos_seleccionados and campos_disponibles_tipo:
        # Seleccionar todos los campos disponibles por defecto
        campos_seleccionados = [campo[0] for campo in campos_disponibles_tipo]
    
    # Preparar el contexto
    context = {
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'tipo_reporte': tipo_reporte,
        'campos_disponibles': campos_disponibles_tipo,
        'campos_seleccionados': campos_seleccionados,
        'tipos_reporte': [
            ('ventas_portabilidad', 'Ventas Portabilidad'),
            ('ventas_prepos', 'Ventas PrePos'),
            ('ventas_upgrade', 'Ventas Upgrade'),
            ('agendamientos', 'Agendamientos'),
            ('comisiones', 'Comisiones'),
            ('escalamientos', 'Escalamientos'),
            ('planes', 'Planes'),
        ],
    }
    
    # Si es una exportación directa a Excel, redirigir
    if accion == 'exportar':
        return redirect('telefonica:reportes_exportar')
    
    # Obtener datos para vista previa (limitados)
    datos = []
    total_registros = 0
    
    try:
        # Convertir fechas de string a objetos datetime con zona horaria
        if fecha_inicio:
            fecha_inicio_naive = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            fecha_inicio_obj = timezone.make_aware(fecha_inicio_naive)
        if fecha_fin:
            fecha_fin_naive = datetime.strptime(fecha_fin, '%Y-%m-%d')
            # Agregar un día para incluir todo el día y establecer hora al final del día
            fecha_fin_naive = fecha_fin_naive.replace(hour=23, minute=59, second=59)
            fecha_fin_obj = timezone.make_aware(fecha_fin_naive)
        
        print(f"Filtrando reportes telefónica desde {fecha_inicio_obj} hasta {fecha_fin_obj}")
        
        if tipo_reporte == 'ventas_portabilidad':
            queryset = VentaPortabilidad.objects.filter(
                fecha_creacion__range=[fecha_inicio_obj, fecha_fin_obj]
            ).select_related('agente', 'backoffice', 'plan_adquiere')
            total_registros = queryset.count()
            datos = list(queryset[:100])  # Limitar para vista previa
            
        elif tipo_reporte == 'ventas_prepos':
            queryset = VentaPrePos.objects.filter(
                fecha_creacion__range=[fecha_inicio_obj, fecha_fin_obj]
            ).select_related('agente', 'cliente_base', 'plan_adquiere')
            total_registros = queryset.count()
            datos = list(queryset[:100])
            
        elif tipo_reporte == 'ventas_upgrade':
            # Usar defer temporalmente para compatibilidad con producción
            queryset = VentaUpgrade.objects.filter(
                fecha_creacion__range=[fecha_inicio_obj, fecha_fin_obj]
            ).select_related('agente', 'cliente_base', 'plan_adquiere').defer('valor_plan_anterior')
            total_registros = queryset.count()
            datos = list(queryset[:100])
            
        elif tipo_reporte == 'agendamientos':
            queryset = Agendamiento.objects.filter(
                fecha_creacion__range=[fecha_inicio_obj, fecha_fin_obj]
            ).select_related('agente')
            total_registros = queryset.count()
            datos = list(queryset[:100])
            
        elif tipo_reporte == 'comisiones':
            queryset = Comision.objects.filter(
                fecha_creacion__range=[fecha_inicio_obj, fecha_fin_obj]
            ).select_related('agente', 'venta')
            total_registros = queryset.count()
            datos = list(queryset[:100])
            
        elif tipo_reporte == 'escalamientos':
            queryset = Escalamiento.objects.filter(
                fecha_escalamiento__range=[fecha_inicio_obj, fecha_fin_obj]
            ).select_related('venta')
            total_registros = queryset.count()
            datos = list(queryset[:100])
            
        elif tipo_reporte == 'planes':
            queryset = Planes_portabilidad.objects.filter(
                fecha_creacion__range=[fecha_inicio_obj, fecha_fin_obj]
            )
            total_registros = queryset.count()
            datos = list(queryset[:100])
        
        # Log de registros solo en desarrollo
        if settings.DEBUG:
            logger = logging.getLogger(__name__)
            logger.info(f"Total de registros encontrados: {total_registros}")
        
    except Exception as e:
        print(f"Error al obtener datos: {e}")
        datos = []
        total_registros = 0
    
    context.update({
        'datos': datos,
        'total_registros': total_registros,
        'mostrando_vista_previa': total_registros > 100,
    })
    
    return render(request, 'telefonica/reportes/reportes.html', context)

@login_required
def exportar_excel(request):
    """Vista para exportar reportes de telefónica a Excel"""
    if request.method == 'POST':
        tipo_reporte = request.POST.get('tipo_reporte', 'ventas_portabilidad')
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        campos_seleccionados = request.POST.getlist('campos')
    else:
        # Si se accede por GET, obtener los parámetros de la URL
        tipo_reporte = request.GET.get('tipo_reporte', 'ventas_portabilidad')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        campos_seleccionados = request.GET.getlist('campos')
        
        # Si no hay parámetros, redirigir a la página de reportes
        if not (tipo_reporte and (fecha_inicio or fecha_fin)):
            return redirect('telefonica:reportes')
    
    print(f"Exportando Excel Telefónica - Tipo: {tipo_reporte}, Fecha inicio: {fecha_inicio}, Fecha fin: {fecha_fin}")
    print(f"Campos seleccionados: {campos_seleccionados}")
    
    # Crear el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Reporte {tipo_reporte.replace('_', ' ').title()}"
    
    # Función auxiliar para obtener valores de campos
    def get_valor(objeto, campo, valor_default=''):
        try:
            # Manejar campos con relaciones (usando __ como separador)
            if '__' in campo:
                partes = campo.split('__')
                obj = objeto
                
                for i, parte in enumerate(partes):
                    if obj is None:
                        return valor_default
                    
                    # Obtener el atributo del objeto actual
                    obj = getattr(obj, parte, None)
                    
                    # Si es None, retornar valor por defecto
                    if obj is None:
                        return valor_default
                
                # Formatear el valor final
                if hasattr(obj, 'strftime'):  # Es una fecha
                    return obj.strftime('%Y-%m-%d %H:%M:%S') if hasattr(obj, 'hour') else obj.strftime('%Y-%m-%d')
                elif isinstance(obj, bool):
                    return 'Sí' if obj else 'No'
                else:
                    return str(obj) if obj is not None else valor_default
            else:
                # Campo directo del objeto
                valor = getattr(objeto, campo, valor_default)
                
                # Formatear según el tipo
                if hasattr(valor, 'strftime'):  # Es una fecha
                    return valor.strftime('%Y-%m-%d %H:%M:%S') if hasattr(valor, 'hour') else valor.strftime('%Y-%m-%d')
                elif isinstance(valor, bool):
                    return 'Sí' if valor else 'No'
                else:
                    return str(valor) if valor is not None else valor_default
                    
        except Exception as e:
            print(f"Error obteniendo valor para campo {campo}: {e}")
            return valor_default
    
    # Obtener datos según el tipo de reporte
    try:
        # Convertir fechas de string a objetos datetime con zona horaria
        if fecha_inicio:
            fecha_inicio_naive = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            fecha_inicio_obj = timezone.make_aware(fecha_inicio_naive)
        if fecha_fin:
            fecha_fin_naive = datetime.strptime(fecha_fin, '%Y-%m-%d')
            # Agregar un día para incluir todo el día y establecer hora al final del día
            fecha_fin_naive = fecha_fin_naive.replace(hour=23, minute=59, second=59)
            fecha_fin_obj = timezone.make_aware(fecha_fin_naive)
        
        if tipo_reporte == 'ventas_portabilidad':
            queryset = VentaPortabilidad.objects.filter(
                fecha_creacion__range=[fecha_inicio_obj, fecha_fin_obj]
            ).select_related('agente', 'backoffice', 'plan_adquiere')
            
        elif tipo_reporte == 'ventas_prepos':
            queryset = VentaPrePos.objects.filter(
                fecha_creacion__range=[fecha_inicio_obj, fecha_fin_obj]
            ).select_related('agente', 'cliente_base', 'plan_adquiere')
            
        elif tipo_reporte == 'ventas_upgrade':
            # Usar defer temporalmente para compatibilidad con producción
            queryset = VentaUpgrade.objects.filter(
                fecha_creacion__range=[fecha_inicio_obj, fecha_fin_obj]
            ).select_related('agente', 'cliente_base', 'plan_adquiere').defer('valor_plan_anterior')
            
        elif tipo_reporte == 'agendamientos':
            queryset = Agendamiento.objects.filter(
                fecha_creacion__range=[fecha_inicio_obj, fecha_fin_obj]
            ).select_related('agente')
            
        elif tipo_reporte == 'comisiones':
            queryset = Comision.objects.filter(
                fecha_creacion__range=[fecha_inicio_obj, fecha_fin_obj]
            ).select_related('agente', 'venta')
            
        elif tipo_reporte == 'escalamientos':
            queryset = Escalamiento.objects.filter(
                fecha_escalamiento__range=[fecha_inicio_obj, fecha_fin_obj]
            ).select_related('venta')
            
        elif tipo_reporte == 'planes':
            queryset = Planes_portabilidad.objects.filter(
                fecha_creacion__range=[fecha_inicio_obj, fecha_fin_obj]
            )
        
        # Si no hay campos seleccionados, usar todos los disponibles
        if not campos_seleccionados:
            campos_seleccionados = [campo[0] for campo in CAMPOS_DISPONIBLES.get(tipo_reporte, [])]
        
        # Escribir encabezados
        campos_dict = dict(CAMPOS_DISPONIBLES.get(tipo_reporte, []))
        for col, campo in enumerate(campos_seleccionados, 1):
            etiqueta = campos_dict.get(campo, campo)
            ws.cell(row=1, column=col, value=etiqueta)
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Escribir datos
        for row, objeto in enumerate(queryset, 2):
            for col, campo in enumerate(campos_seleccionados, 1):
                valor = get_valor(objeto, campo)
                ws.cell(row=row, column=col, value=valor)
        
        # Ajustar ancho de columnas
        for col in range(1, len(campos_seleccionados) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 20
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Nombre del archivo
        fecha_actual = timezone.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f"reporte_telefonica_{tipo_reporte}_{fecha_actual}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        
        # Guardar el libro en la respuesta
        wb.save(response)
        
        return response
        
    except Exception as e:
        print(f"Error al generar Excel: {e}")
        return redirect('telefonica:reportes')